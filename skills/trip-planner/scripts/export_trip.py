#!/usr/bin/env python3
"""Export a trip-planner HTML file to XLSX and PDF without the user clicking buttons.

The HTML files this skill produces contain two export buttons (XLSX via SheetJS,
PDF via window.print) that work for humans but don't run end-to-end via the MCP.
This script provides a deterministic alternative: parse the table rows out of
the HTML, write a real .xlsx via openpyxl, and render a real .pdf via
chrome --headless --print-to-pdf.

Usage:
    python3 export_trip.py --html ~/Desktop/trip_xxx.html --out-dir ~/Desktop/
    python3 export_trip.py --html trip.html --out-dir ./out --skip-pdf
    python3 export_trip.py --html trip.html --out-dir ./out --skip-xlsx

Requirements:
    pip install openpyxl beautifulsoup4
    Chrome / Chromium available on PATH (or pass --chrome /path/to/chrome)
"""
from __future__ import annotations

import argparse
import html as _html
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

try:
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("Missing dependency: pip install beautifulsoup4")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


ROW_COLOURS = {
    "type-flight": "E8F1FD",
    "type-hotel": "E8F8ED",
    "type-transfer": "FFF3E0",
    "type-tbd": "ECECEC",
}


TYPE_LABEL = {"flight": "Перелёт", "hotel": "Отель", "transfer": "Трансфер"}


def parse_trip_html(html_path: Path) -> tuple[str, list[dict]]:
    """Return (title, rows). Prefers the structured `<script id="trip-data">`
    JSON single-source-of-truth block; falls back to scraping the rendered
    table for outputs generated before the JSON-SoT refactor."""
    text = html_path.read_text(encoding="utf-8")
    m = re.search(r'<script id="trip-data" type="application/json"[^>]*>(.*?)</script>',
                  text, re.DOTALL | re.IGNORECASE)
    if m:
        try:
            return _rows_from_json(json.loads(m.group(1)))
        except (json.JSONDecodeError, ValueError):
            pass  # malformed → fall back to scraping the table
    return _rows_from_table(text)


def _flat(s: str) -> str:
    return _html.unescape(re.sub(r"<[^>]+>", " ", s or "")).replace("\n", " ").strip()


def _rows_from_json(data: dict) -> tuple[str, list[dict]]:
    title = _html.unescape((data.get("meta", {}) or {}).get("title") or "Trip").strip()
    rows = []
    for i, r in enumerate(data.get("rows", []), 1):
        rating = "—"
        if r.get("rating"):
            rt = r["rating"]
            rating = f"TA {rt.get('ta', '')}/5, {_flat(rt.get('taReviews', ''))}, Ostrovok {rt.get('ostrovok', '')}"
        date = _flat(r.get("date", "")) + (f" ({_flat(r.get('dateNote'))})" if r.get("dateNote") else "")
        desc = _flat(r.get("title", "")) + (f" — {_flat(r.get('sub'))}" if r.get("sub") else "")
        time = _flat(r.get("time", "")) + (f" ({_flat(r.get('timeNote'))})" if r.get("timeNote") else "")
        details = _flat(r.get("details", "")) + (f" {_flat(r.get('detailsNote'))}" if r.get("detailsNote") else "")
        cells = [str(i), TYPE_LABEL.get(r.get("type"), r.get("type", "")), date, desc,
                 time.strip(), details.strip(), rating, _flat(r.get("price", "")) or "—", ""]
        links = [l.get("url", "") for l in (r.get("links") or [])]
        rows.append({"kind": f"type-{r.get('type')}", "cells": cells, "links": links})
    return title, rows


def _rows_from_table(text: str) -> tuple[str, list[dict]]:
    soup = BeautifulSoup(text, "html.parser")
    title = (soup.title.string or "Trip").strip() if soup.title else "Trip"

    table = soup.find("table", id="tripTable")
    if not table:
        raise RuntimeError("No <table id='tripTable'> found — is this a trip-planner HTML?")

    rows = []
    for tr in table.find("tbody").find_all("tr"):
        row_class = " ".join(tr.get("class", []))
        kind = next((k for k in ROW_COLOURS if k in row_class), None)
        cells = [td.get_text(" ", strip=True) for td in tr.find_all("td")]
        if not cells:
            continue
        links = [a.get("href", "") for a in tr.find_all("a", href=True)]
        rows.append({"kind": kind, "cells": cells, "links": links})

    return title, rows


def write_xlsx(title: str, rows: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Маршрут"

    headers = ["#", "Тип", "Дата", "Описание", "Время / Заезд", "Детали",
               "Рейтинг", "Цена (2 чел.)", "Ссылки"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FAFAFA")

    for row in rows:
        cells = list(row["cells"])
        # Pad to header length
        while len(cells) < len(headers):
            cells.append("")
        # Replace the empty Links cell with a joined link list
        if row["links"]:
            cells[-1] = " | ".join(row["links"])
        ws.append(cells)
        excel_row = ws.max_row
        colour = ROW_COLOURS.get(row["kind"])
        if colour:
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = PatternFill("solid", fgColor=colour)
        for col in range(1, len(headers) + 1):
            ws.cell(row=excel_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    column_widths = [4, 12, 18, 28, 22, 35, 18, 16, 60]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def render_pdf(html_path: Path, out_path: Path, chrome_path: str | None) -> None:
    chrome = chrome_path or _find_chrome()
    if not chrome:
        raise RuntimeError(
            "Chrome / Chromium not found. Install Chrome or pass --chrome /path/to/chrome"
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    abs_html = html_path.resolve()
    cmd = [
        chrome,
        "--headless=new",
        "--disable-gpu",
        "--no-sandbox",
        f"--print-to-pdf={out_path}",
        "--print-to-pdf-no-header",
        f"file://{abs_html}",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if not out_path.exists():
        raise RuntimeError(f"chrome failed to produce PDF.\nstdout: {result.stdout}\nstderr: {result.stderr}")


def _find_chrome() -> str | None:
    candidates = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Chromium.app/Contents/MacOS/Chromium",
        shutil.which("google-chrome"),
        shutil.which("chromium"),
        shutil.which("chrome"),
    ]
    for c in candidates:
        if c and Path(c).exists():
            return c
    return None


def _slug(s: str) -> str:
    s = re.sub(r"[^\w\s.-]", "", s, flags=re.UNICODE)
    s = re.sub(r"\s+", "_", s).strip("_")
    return s or "trip"


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--html", required=True, type=Path, help="Path to the trip HTML")
    p.add_argument("--out-dir", required=True, type=Path, help="Where to write the .xlsx / .pdf")
    p.add_argument("--basename", help="Override output base name (default: derived from <title>)")
    p.add_argument("--skip-xlsx", action="store_true")
    p.add_argument("--skip-pdf", action="store_true")
    p.add_argument("--chrome", help="Explicit Chrome/Chromium binary path")
    args = p.parse_args()

    html_path = args.html.expanduser()
    out_dir = args.out_dir.expanduser()
    if not html_path.exists():
        sys.exit(f"Not found: {html_path}")

    title, rows = parse_trip_html(html_path)
    base = args.basename or _slug(title)
    out_dir.mkdir(parents=True, exist_ok=True)

    produced = []
    if not args.skip_xlsx:
        xlsx_path = out_dir / f"{base}.xlsx"
        write_xlsx(title, rows, xlsx_path)
        produced.append(xlsx_path)

    if not args.skip_pdf:
        pdf_path = out_dir / f"{base}.pdf"
        render_pdf(html_path, pdf_path, args.chrome)
        produced.append(pdf_path)

    for path in produced:
        print(path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
